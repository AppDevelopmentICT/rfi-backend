import os
from openpyxl import load_workbook
from openpyxl.styles import Alignment

from app.config import SOURCE_FILE, FILES_DIR
from app.services.external.ollama import ask_ollama


def _load_workbook():
    """Load the source Excel file, raising FileNotFoundError if missing."""
    if not os.path.exists(SOURCE_FILE):
        raise FileNotFoundError(f"Source file not found: {SOURCE_FILE}")
    return load_workbook(SOURCE_FILE)


def read_all_sheets() -> dict:
    """Read all sheets from the Excel file and return headers + data."""
    wb = _load_workbook()
    sheets = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            sheets[sheet_name] = {"headers": [], "data": []}
            continue

        headers = [
            str(h) if h is not None else f"col_{i}"
            for i, h in enumerate(all_rows[0])
        ]
        data = []
        for row in all_rows[1:]:
            row_dict = {}
            for idx, val in enumerate(row):
                key = headers[idx] if idx < len(headers) else f"col_{idx}"
                row_dict[key] = val
            data.append(row_dict)

        sheets[sheet_name] = {"headers": headers, "data": data}

    wb.close()
    return sheets


async def auto_fill_sheets() -> dict:
    """Fill every empty cell in every sheet using the LLM."""
    wb = _load_workbook()
    results = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        all_rows = list(ws.iter_rows(values_only=True))
        if len(all_rows) < 2:
            continue

        headers = [
            str(h).strip() if h is not None else "" for h in all_rows[0]
        ]


        question_col = None
        for i, h in enumerate(headers):
            if "question" in h.lower():
                question_col = i
                break

        if question_col is None:
            continue


        fill_cols = []
        for i, h in enumerate(headers):
            if i == question_col:
                continue
            low = h.lower()
            if low in ("no", "no.", "n"):
                continue
            fill_cols.append(i)

        if not fill_cols:
            continue


        for row_idx, row in enumerate(all_rows[1:], start=2):
            question = row[question_col]
            if not question or str(question).strip() == "":
                continue

            for col_idx in fill_cols:
                existing = row[col_idx] if col_idx < len(row) else None
                if existing is not None and str(existing).strip() != "":
                    continue

                col_header = headers[col_idx]
                answer = await ask_ollama(str(question).strip(), col_header)

                cell = ws.cell(
                    row=row_idx, column=col_idx + 1, value=answer
                )
                cell.alignment = Alignment(wrap_text=True, vertical="top")

                results.append(
                    {
                        "sheet": sheet_name,
                        "row": row_idx,
                        "column": col_header,
                        "question": str(question).strip(),
                        "answer": answer,
                    }
                )

    output_path = os.path.join(FILES_DIR, "RFI-Examples_answered.xlsx")
    wb.save(output_path)
    wb.close()

    return {
        "message": f"Filled {len(results)} cells",
        "output_file": "RFI-Examples_answered.xlsx",
        "results": results,
    }
