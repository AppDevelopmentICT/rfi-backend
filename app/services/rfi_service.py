import io
from typing import Optional

from openpyxl import load_workbook
from openpyxl.styles import Alignment

from app.services.ollama_service import ask_ollama


def parse_excel_bytes(file_bytes: bytes) -> dict:
    """Parse uploaded Excel bytes and return all sheets with headers + data."""
    wb = load_workbook(io.BytesIO(file_bytes))
    sheets = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            sheets[sheet_name] = {"headers": [], "data": []}
            continue

        headers = [
            str(h) if h is not None else f"col_{i}"
            for i, h in enumerate(all_rows[0])
        ]
        data = []
        for row in all_rows[1:]:
            row_dict = {}
            for idx, val in enumerate(row):
                key = headers[idx] if idx < len(headers) else f"col_{idx}"
                row_dict[key] = val
            data.append(row_dict)

        sheets[sheet_name] = {"headers": headers, "data": data}

    wb.close()
    return sheets


def _classify_columns(
    headers: list[str],
    data_rows: list[tuple],
    context_columns: Optional[list[str]] = None,
    fill_columns: Optional[list[str]] = None,
) -> tuple[list[int], list[int]]:
    """Classify columns as context (input) or fill (to be filled by LLM).

    Auto-detection logic (when not overridden):
    - A column is "context" if >50% of its data rows are non-empty.
    - A column is "fill" if it has at least one empty cell in data rows.
    - Columns that are 100% empty across all rows are also "fill" targets.

    Args:
        headers:         List of column header strings.
        data_rows:       All data rows (excluding the header row).
        context_columns: Optional user-specified context column names.
        fill_columns:    Optional user-specified fill column names.

    Returns:
        (context_col_indices, fill_col_indices)
    """
    num_rows = len(data_rows)
    if num_rows == 0:
        return [], []

    # ── user-specified overrides ────────────────────────────────────
    header_lower_map = {h.strip().lower(): i for i, h in enumerate(headers)}

    if context_columns and fill_columns:
        ctx = [header_lower_map[c.strip().lower()] for c in context_columns if c.strip().lower() in header_lower_map]
        fill = [header_lower_map[c.strip().lower()] for c in fill_columns if c.strip().lower() in header_lower_map]
        return ctx, fill

    # ── auto-detect ─────────────────────────────────────────────────
    fill_count = {}   # col_idx -> number of empty cells
    filled_count = {} # col_idx -> number of non-empty cells

    for col_idx in range(len(headers)):
        empty = 0
        filled = 0
        for row in data_rows:
            val = row[col_idx] if col_idx < len(row) else None
            if val is None or str(val).strip() == "":
                empty += 1
            else:
                filled += 1
        fill_count[col_idx] = empty
        filled_count[col_idx] = filled

    ctx_indices = []
    fill_indices = []

    for col_idx in range(len(headers)):
        has_empties = fill_count[col_idx] > 0
        fill_ratio = filled_count[col_idx] / num_rows if num_rows else 0

        if has_empties:
            fill_indices.append(col_idx)
            # also treat as context if it has substantial data (>50%)
            if fill_ratio > 0.5:
                ctx_indices.append(col_idx)
        else:
            # fully filled — context only
            ctx_indices.append(col_idx)

    # if user specified only context_columns, auto-detect fill
    if context_columns:
        ctx_indices = [header_lower_map[c.strip().lower()] for c in context_columns if c.strip().lower() in header_lower_map]
    # if user specified only fill_columns, auto-detect context
    if fill_columns:
        fill_indices = [header_lower_map[c.strip().lower()] for c in fill_columns if c.strip().lower() in header_lower_map]

    return ctx_indices, fill_indices


def _build_context_prompt(headers: list[str], row: tuple, context_indices: list[int]) -> str:
    """Build a context string from all context columns for a given row."""
    parts = []
    for idx in context_indices:
        val = row[idx] if idx < len(row) else None
        if val is not None and str(val).strip():
            parts.append(f"{headers[idx]}: {str(val).strip()}")
    return "\n".join(parts)


async def auto_fill_bytes(
    file_bytes: bytes,
    model: Optional[str] = None,
    context_columns: Optional[list[str]] = None,
    fill_columns: Optional[list[str]] = None,
) -> dict:
    """Fill empty cells in an uploaded Excel file using the LLM.

    Column detection is fully dynamic:
    - Context columns (used as input to LLM) = columns where >50% rows have data
    - Fill columns (LLM will generate answers) = columns with empty cells

    The user can optionally override with `context_columns` and `fill_columns`.

    Args:
        file_bytes:      Raw bytes of the uploaded .xlsx file.
        model:           Optional Ollama model name; falls back to config default.
        context_columns: Optional list of column names to use as LLM context.
        fill_columns:    Optional list of column names the LLM should fill.

    Returns:
        Dict with ``filled_bytes``, ``message``, ``results``, and ``column_info``.
    """
    wb = load_workbook(io.BytesIO(file_bytes))
    results = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        all_rows = list(ws.iter_rows(values_only=True))
        if len(all_rows) < 2:
            continue

        headers = [
            str(h).strip() if h is not None else f"col_{i}"
            for i, h in enumerate(all_rows[0])
        ]
        data_rows = all_rows[1:]

        ctx_indices, fill_indices = _classify_columns(
            headers, data_rows, context_columns, fill_columns
        )

        if not ctx_indices or not fill_indices:
            continue

        # iterate rows and fill empty cells
        for row_idx, row in enumerate(data_rows, start=2):
            context_prompt = _build_context_prompt(headers, row, ctx_indices)
            if not context_prompt.strip():
                continue  # no context data in this row, skip

            for col_idx in fill_indices:
                existing = row[col_idx] if col_idx < len(row) else None
                if existing is not None and str(existing).strip() != "":
                    continue  # already has data, skip

                col_header = headers[col_idx]
                answer = await ask_ollama(
                    context_prompt, col_header, model=model
                )

                cell = ws.cell(
                    row=row_idx, column=col_idx + 1, value=answer
                )
                cell.alignment = Alignment(wrap_text=True, vertical="top")

                results.append(
                    {
                        "sheet": sheet_name,
                        "row": row_idx,
                        "column": col_header,
                        "context": context_prompt,
                        "answer": answer,
                    }
                )

    # save to in-memory buffer
    output_buf = io.BytesIO()
    wb.save(output_buf)
    wb.close()
    output_buf.seek(0)

    return {
        "filled_bytes": output_buf.getvalue(),
        "message": f"Filled {len(results)} cells",
        "results": results,
    }

